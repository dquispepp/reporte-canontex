"""
Reporte diario de ventas Canontex (OMSWeb).
Ejecuta login, consulta el dashboard de ventas y genera un Excel diario.

Uso:
    python reporte_diario.py                            # reporta el día de ayer
    python reporte_diario.py 2026-07-10                 # reporta una fecha específica
    python reporte_diario.py 2026-07-01 2026-07-10       # reporta un rango de fechas
"""
import os
import sys
import re
import requests
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BASE_URL = "https://canontex.bbr.cl"
USERNAME = os.environ["OMS_USER"]
PASSWORD = os.environ["OMS_PASS"]
OUTPUT_DIR = os.environ.get("REPORTE_OUTPUT_DIR", "./reportes")


def login(session: requests.Session) -> None:
    login_page = session.get(f"{BASE_URL}/webapp/login?logout")
    login_page.raise_for_status()
    match = re.search(r'name="_csrf" value="([^"]+)"', login_page.text)
    if not match:
        raise RuntimeError("No se encontró el token _csrf en la página de login")
    csrf = match.group(1)

    resp = session.post(
        f"{BASE_URL}/login",
        data={"_csrf": csrf, "username": USERNAME, "password": PASSWORD},
    )
    resp.raise_for_status()
    if "Dashboard" not in resp.text and "location.href" not in resp.text:
        raise RuntimeError("Login fallido: credenciales inválidas o CSRF vencido")


def get_dashboard_venta(session: requests.Session, fecha_ini: datetime, fecha_fin: datetime, canal: str = "") -> dict:
    params = {
        "fechaIni": fecha_ini.strftime("%Y-%-m-%-d 00:00:00") if os.name != "nt" else f"{fecha_ini.year}-{fecha_ini.month}-{fecha_ini.day} 00:00:00",
        "fechaFin": fecha_fin.strftime("%Y-%-m-%-d 00:00:00") if os.name != "nt" else f"{fecha_fin.year}-{fecha_fin.month}-{fecha_fin.day} 00:00:00",
        "canal": canal,
    }
    resp = session.get(f"{BASE_URL}/webapp/wsrest/getDashboardVenta", params=params)
    resp.raise_for_status()
    if not resp.text.strip():
        return {}
    return resp.json()


def build_report(data: dict, fecha_inicio: datetime, fecha_fin: datetime, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    if fecha_inicio.date() == fecha_fin.date():
        periodo = fecha_inicio.strftime("%d-%m-%Y")
    else:
        periodo = f"{fecha_inicio.strftime('%d-%m-%Y')} a {fecha_fin.strftime('%d-%m-%Y')}"
    ws["A1"] = f"Reporte de ventas - Canontex - {periodo}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    resumen = [
        ("Cantidad de ventas", data.get("cantidadVentas", 0)),
        ("Cantidad de unidades", data.get("cantidadUnidades", 0)),
        ("Total ventas (CLP)", data.get("totalVentas", 0)),
        ("Venta promedio (CLP)", data.get("ventaPromedio", 0)),
    ]
    ws["A3"] = "Indicador"
    ws["B3"] = "Valor"
    for cell in ("A3", "B3"):
        ws[cell].font = header_font
        ws[cell].fill = header_fill
    for i, (label, value) in enumerate(resumen, start=4):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value

    row = 10
    ws[f"A{row}"] = "Ventas por canal"
    ws[f"A{row}"].font = bold
    row += 1
    ws[f"A{row}"] = "Canal"
    ws[f"B{row}"] = "Total (CLP)"
    ws[f"A{row}"].font = header_font
    ws[f"B{row}"].font = header_font
    ws[f"A{row}"].fill = header_fill
    ws[f"B{row}"].fill = header_fill
    row += 1
    canales = {
        "ECOMMERCE": data.get("graficoTotalEcommerce", []),
        "RETAIL (B2B)": data.get("graficoTotalB2B", []),
        "KIOSCO": data.get("graficoTotalKiosco", []),
        "MKP": data.get("graficoTotalMKP", []),
    }
    for nombre, serie in canales.items():
        total = sum(item.get("cantidad", 0) for item in serie)
        ws[f"A{row}"] = nombre
        ws[f"B{row}"] = total
        row += 1

    row += 2
    ws[f"A{row}"] = "Top productos (SKU)"
    ws[f"A{row}"].font = bold
    row += 1
    headers = ["SKU", "Nombre", "Cantidad", "Valor (CLP)", "% del total"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    row += 1
    for item in data.get("tablaSKU", []):
        ws.cell(row=row, column=1, value=item.get("titulo"))
        ws.cell(row=row, column=2, value=item.get("nombre"))
        ws.cell(row=row, column=3, value=item.get("cantidad"))
        ws.cell(row=row, column=4, value=item.get("valor"))
        ws.cell(row=row, column=5, value=item.get("porcentaje"))
        row += 1

    row += 2
    ws[f"A{row}"] = "Ventas por marca"
    ws[f"A{row}"].font = bold
    row += 1
    headers = ["Marca", "Cantidad", "Valor (CLP)", "% del total"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    row += 1
    for item in data.get("tablaMarcas", []):
        ws.cell(row=row, column=1, value=item.get("titulo"))
        ws.cell(row=row, column=2, value=item.get("cantidad"))
        ws.cell(row=row, column=3, value=item.get("valor"))
        ws.cell(row=row, column=4, value=item.get("porcentaje"))
        row += 1

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)


def main():
    if len(sys.argv) > 1:
        fecha_inicio = datetime.strptime(sys.argv[1], "%Y-%m-%d")
        fecha_fin = datetime.strptime(sys.argv[2], "%Y-%m-%d") if len(sys.argv) > 2 else fecha_inicio
    else:
        fecha_inicio = fecha_fin = datetime.now() - timedelta(days=1)

    session = requests.Session()
    login(session)
    data = get_dashboard_venta(session, fecha_inicio, fecha_fin)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if fecha_inicio.date() == fecha_fin.date():
        filename = f"reporte_ventas_{fecha_inicio.strftime('%Y-%m-%d')}.xlsx"
    else:
        filename = f"reporte_ventas_{fecha_inicio.strftime('%Y-%m-%d')}_a_{fecha_fin.strftime('%Y-%m-%d')}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, filename)
    build_report(data, fecha_inicio, fecha_fin, output_path)
    print(f"Reporte generado: {output_path}")


if __name__ == "__main__":
    main()
