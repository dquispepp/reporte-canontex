"""Genera el Excel ejecutivo diario (3 hojas) con solo los pedidos atrasados a HOY."""
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
HDR_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=13)
RED_STRONG = PatternFill("solid", fgColor="E06666")  # > 30 días
RED = PatternFill("solid", fgColor="F4CCCC")          # > 1 semana
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREEN = PatternFill("solid", fgColor="D9EAD3")
THIN = Border(*[Side(style="thin", color="CCCCCC")] * 4)

RETIRO = ("Retiro en Tienda", "Cross Docking")
DOMICILIO = ("Despacho a Domicilio", "Fecha Pactada")

GLOSARIO = [
    ("TIENDA CONFIRMAR ENTREGA", "La tienda debe confirmar si el cliente ya retiró el pedido."),
    ("GESTIÓN PENDIENTE — REVISAR WMS", "El pedido no figura despachado en WMS; revisar por qué."),
    ("NO EXISTE EN WMS", "El pedido está en el gestor pero no ingresó al WMS; investigar."),
    ("SAC DEVOLUCIÓN DINERO", "Cliente no retiró/venció plazo; SAC gestiona devolución."),
    ("BODEGA WMS", "Aún no despachado; empujar la operación en bodega."),
    ("ALERTA ETIQUETA", "Bodega despachó pero el courier no lo tomó (posible etiqueta mal pegada)."),
    ("COURIER PICKUP", "El courier no ha retirado el pedido del andén."),
    ("COURIER RUTA", "Atraso dentro de la red del courier."),
    ("COURIER FALLIDO", "Entrega fallida; reprogramar."),
    ("RECONTACTAR CLIENTE", "Coordinar nueva entrega con el cliente."),
]


def _fmt(v):
    if pd.isna(v):
        return ""
    if isinstance(v, pd.Timestamp):
        return v.strftime("%d-%m-%Y")
    return v


def _style_sheet(ws, n_cols, dias_col_idx=None, header_row=1):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_cols)}{header_row}"
    ws.freeze_panes = f"A{header_row + 1}"
    for c in range(1, n_cols + 1):
        letter = get_column_letter(c)
        width = max(
            [len(str(ws.cell(row=r, column=c).value or "")) for r in range(header_row, min(ws.max_row, 200) + 1)]
            + [10]
        )
        ws.column_dimensions[letter].width = min(width + 2, 40)
    if dias_col_idx and ws.max_row > header_row:
        col = get_column_letter(dias_col_idx)
        rng = f"{col}{header_row + 1}:{col}{ws.max_row}"
        # Semáforo: rojo fuerte > 30 días, rojo claro > 1 semana (8-30 días)
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["30"], fill=RED_STRONG))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["8", "30"], fill=RED))


def _write_table(ws, df, columnas, dias_label="Días de atraso", start_row=1):
    labels = [c[1] for c in columnas]
    getters = [c[0] for c in columnas]
    for j, lab in enumerate(labels, start=1):
        ws.cell(row=start_row, column=j, value=lab)
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, g in enumerate(getters, start=1):
            ws.cell(row=i, column=j, value=_fmt(g(row) if callable(g) else row.get(g)))
    dias_idx = labels.index(dias_label) + 1 if dias_label in labels else None
    _style_sheet(ws, len(labels), dias_idx, header_row=start_row)
    return start_row + 1 + len(df)


def _hoja_resumen(wb, atrasados, kpis):
    ws = wb.active
    ws.title = "Resumen Ejecutivo"
    ws["A1"] = f"Resumen de pedidos atrasados — {kpis['fecha']}"
    ws["A1"].font = TITLE_FONT

    # matriz Canal x Tipo de Despacho
    r0 = 3
    ws.cell(row=r0, column=1, value="Atrasados por Canal × Tipo de Despacho (count / días prom.)").font = HDR_FONT
    r0 += 1
    if len(atrasados):
        cnt = atrasados.pivot_table(index="canal", columns="tipo_despacho", values="dias_atraso",
                                    aggfunc="count", margins=True, margins_name="Total")
        prom = atrasados.pivot_table(index="canal", columns="tipo_despacho", values="dias_atraso",
                                     aggfunc="mean", margins=True, margins_name="Total")
        cols = list(cnt.columns)
        ws.cell(row=r0, column=1, value="Canal").font = HDR_FONT
        for j, c in enumerate(cols, start=2):
            ws.cell(row=r0, column=j, value=c).font = HDR_FONT
        for i, canal in enumerate(cnt.index, start=r0 + 1):
            ws.cell(row=i, column=1, value=canal)
            for j, c in enumerate(cols, start=2):
                n = cnt.loc[canal, c]
                p = prom.loc[canal, c]
                val = "" if pd.isna(n) else f"{int(n)} / {0 if pd.isna(p) else round(p, 1)}"
                ws.cell(row=i, column=j, value=val)
        r_kpi = r0 + len(cnt.index) + 2
    else:
        ws.cell(row=r0, column=1, value="(sin atrasados)")
        r_kpi = r0 + 2

    ws.cell(row=r_kpi, column=1, value="KPIs").font = HDR_FONT
    kpi_rows = [
        ("Total pedidos atrasados", kpis["atrasados"]),
        ("Pedidos NO creados en WMS", kpis["no_wms"]),
        ("Facturas reserva (13-) canceladas sin OV (17-)", kpis["reserva_sin_ov"]),
        ("% OTIF (on-time al cliente)", f"{kpis['pct_otif']}%"),
        ("Días de atraso promedio", kpis["dias_atraso_prom"]),
        ("Total pedidos abiertos", kpis["total_abiertos"]),
    ]
    for k, (lab, val) in enumerate(kpi_rows, start=r_kpi + 1):
        ws.cell(row=k, column=1, value=lab)
        ws.cell(row=k, column=2, value=val)

    r_diag = r_kpi + len(kpi_rows) + 2
    ws.cell(row=r_diag, column=1, value="Breakdown por diagnóstico de causa raíz").font = HDR_FONT
    for k, (diag, n) in enumerate(kpis["por_diagnostico"], start=r_diag + 1):
        ws.cell(row=k, column=1, value=diag)
        ws.cell(row=k, column=2, value=int(n))

    ws.column_dimensions["A"].width = 50
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16


def _hoja_glosario(ws, start_row, n_cols):
    r = start_row + 1
    ws.cell(row=r, column=1, value="Glosario").font = HDR_FONT
    for k, (termino, desc) in enumerate(GLOSARIO, start=r + 1):
        ws.cell(row=k, column=1, value=termino).font = Font(bold=True)
        ws.cell(row=k, column=2, value=desc)


def _hoja_retiro(wb, atrasados):
    ws = wb.create_sheet("Retiro Tienda y CrossDocking")
    df = atrasados[atrasados["tipo_despacho"].isin(RETIRO)].copy()
    columnas = [
        ("num_envio", "N° de Envío"), ("orden_compra", "Orden de Compra"),
        ("cliente_nombre", "Nombre Cliente"),
        ("canal", "Canal de Venta"), ("familia", "Tipo de Envío"),
        ("tienda_retiro", "Tienda de Retiro"),
        ("fecha_trx", "Fecha Trx"), ("fecha_entrega", "Fecha de Entrega"),
        ("dias_atraso", "Días de atraso"), ("metodo_pago", "Método de Pago"),
        ("transportista_norm", "Transportista"), ("estado", "Estado Reporte"),
        ("estado_wms", "Estado WMS"), ("sap_quiebre", "Quiebre SAP"),
        ("n_doc", "N° Doc."),
        ("marca_gestion", "Marca Gestión"),
        ("obs_retiro", "Observación"),
    ]
    next_row = _write_table(ws, df, columnas)
    _hoja_glosario(ws, next_row, len(columnas))


def _hoja_domicilio(wb, atrasados):
    ws = wb.create_sheet("Domicilio y Fecha Pactada")
    df = atrasados[atrasados["tipo_despacho"].isin(DOMICILIO)].copy()
    columnas = [
        ("num_envio", "N° de Envío"), ("orden_compra", "Orden de Compra"),
        ("cliente_nombre", "Nombre Cliente"),
        ("canal", "Canal de Venta"), ("familia", "Tipo de Envío"),
        ("region", "Región"), ("comuna", "Comuna"),
        ("fecha_trx", "Fecha Trx"), ("fecha_entrega", "Fecha de Entrega"),
        ("fecha_despacho_esperada", "Fecha Despacho Esperada"),
        ("dias_atraso", "Días de atraso"), ("transportista_norm", "Transportista"),
        ("tracking", "N° Tracking"), ("estado", "Estado Reporte"),
        ("estado_wms", "Estado WMS"), ("fecha_despacho", "FechaDespacho WMS"),
        ("sap_quiebre", "Quiebre SAP"), ("n_doc", "N° Doc."),
        ("metodo_pago", "Método de Pago"), ("responsable", "Responsable"),
        ("obs_domicilio", "Observación"),
    ]
    next_row = _write_table(ws, df, columnas)
    _hoja_glosario(ws, next_row, len(columnas))


def _hoja_contacto(wb, atrasados):
    ws = wb.create_sheet("Contacto")
    df = atrasados.copy()
    columnas = [
        ("num_envio", "N° de Envío"), ("orden_compra", "Orden de Compra"),
        ("cliente_nombre", "Nombre Cliente"), ("telefono", "Teléfono"),
        ("correo", "Correo"), ("canal", "Canal de Venta"),
        ("tipo_despacho", "Tipo de Despacho"), ("dias_atraso", "Días de atraso"),
        ("diagnostico", "Diagnóstico"),
    ]
    _write_table(ws, df, columnas)


def _sla_label(v):
    if pd.isna(v):
        return "-"
    return "Cumple" if bool(v) else "No cumple"


def _hoja_sla_detalle(wb, resultado):
    """Hoja 'SLA Detalle': todos los pedidos con marca de cumplimiento por SLA."""
    ws = wb.create_sheet("SLA Detalle")
    df = resultado.copy()
    df["_sla_ecom_lbl"] = df["sla_ecommerce"].apply(_sla_label)
    df["_sla_oper_lbl"] = df["sla_operacion"].apply(_sla_label)
    df["_sla_cour_lbl"] = df["sla_courier"].apply(_sla_label)
    columnas = [
        ("num_envio", "N° de Envío"), ("orden_compra", "Orden de Compra"),
        ("canal", "Canal"), ("tipo_despacho", "Tipo Despacho"),
        ("familia", "Familia"), ("region", "Región"),
        ("fecha_trx", "Fecha Trx"), ("fecha_creacion", "Fecha Creación WMS"),
        ("fecha_despacho", "Fecha Despacho"), ("fecha_compromiso", "Fecha Compromiso"),
        ("fecha_entrega_real", "Fecha Entrega Real"),
        ("estado", "Estado OMS"), ("estado_wms", "Estado WMS"),
        ("transportista_norm", "Transportista"),
        ("_sla_ecom_lbl", "SLA Ecommerce (≥72h creación→compromiso)"),
        ("_sla_oper_lbl", "SLA Operación (despacho→compromiso)"),
        ("_sla_cour_lbl", "SLA Courier (última milla)"),
        ("atrasado", "Atrasado hoy"), ("dias_atraso", "Días de atraso"),
        ("diagnostico", "Diagnóstico"),
    ]
    _write_table(ws, df, columnas, dias_label="Días de atraso")

    # Formato condicional verde/rojo para las 3 columnas SLA
    if ws.max_row > 1:
        labels = [c[1] for c in columnas]
        for lbl in ("SLA Ecommerce (≥72h creación→compromiso)",
                    "SLA Operación (despacho→compromiso)",
                    "SLA Courier (última milla)"):
            col_idx = labels.index(lbl) + 1
            letter = get_column_letter(col_idx)
            rng = f"{letter}2:{letter}{ws.max_row}"
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Cumple"'], fill=GREEN))
            ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"No cumple"'], fill=RED))


def build_excel(resultado: pd.DataFrame, kpis: dict, output_path) -> None:
    atrasados = resultado[resultado["atrasado"]].copy()
    wb = Workbook()
    _hoja_resumen(wb, atrasados, kpis)
    _hoja_retiro(wb, atrasados)
    _hoja_domicilio(wb, atrasados)
    _hoja_contacto(wb, atrasados)
    _hoja_sla_detalle(wb, resultado)
    wb.save(output_path)
